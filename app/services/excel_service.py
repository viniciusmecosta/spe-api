import os
import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload

from app.core.config import settings
from app.domain.models.user import User
from app.repositories.company_repository import company_repository
from app.services.report_service import report_service
from app.utils.formatters import format_short_name

FONT_NAME = "Times New Roman"
TIME_FORMAT = "[h]:mm"
NOT_REGISTERED = "Não registrado"


class ExcelService:
    def __init__(self):
        self._setup_styles()

    def _setup_styles(self):
        self.font_regular = Font(name=FONT_NAME, size=11, color="000000")
        self.font_bold = Font(name=FONT_NAME, size=11, bold=True, color="000000")
        self.font_italic = Font(name=FONT_NAME, size=11, italic=True, color="64748B")

        self.font_title_large = Font(name=FONT_NAME, size=16, bold=True, color="000000")
        self.font_subtitle = Font(name=FONT_NAME, size=12, bold=True, color="000000")

        self.font_key = Font(name=FONT_NAME, size=11, bold=True, color="000000")
        self.font_val = Font(name=FONT_NAME, size=11, color="000000")

        self.header_font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="475569", end_color="475569", fill_type="solid")
        
        thin = Side(style='thin', color="CBD5E1")
        self.border_standard = Border(left=thin, right=thin, top=thin, bottom=thin)
        self.border_top_bottom = Border(top=thin, bottom=thin)
        self.border_left_top_bottom = Border(left=thin, top=thin, bottom=thin)
        self.border_right_top_bottom = Border(right=thin, top=thin, bottom=thin)

        self.fill_weekend = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        self.fill_holiday = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        self.fill_absence = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        self.fill_excused = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        self.fill_section_title = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

        self.font_absence = Font(name=FONT_NAME, size=11, bold=False, color="000000")
        self.font_excused = Font(name=FONT_NAME, size=11, bold=False, color="000000")
        self.font_holiday = Font(name=FONT_NAME, size=11, bold=False, color="000000")
        self.font_weekend = Font(name=FONT_NAME, size=11, bold=False, color="000000")

        self.align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        self.align_right = Alignment(horizontal='right', vertical='center', wrap_text=True)
        
        self.MAX_COLS = 24
        self.COL_WIDTH = 5

    def _set_columns_width(self, ws):
        for col in range(1, self.MAX_COLS + 1):
            ws.column_dimensions[get_column_letter(col)].width = self.COL_WIDTH

    def _format_cnpj(self, cnpj: str) -> str:
        if not cnpj: return NOT_REGISTERED
        c = re.sub(r'\D', '', cnpj)
        if len(c) == 14:
            return f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}"
        return cnpj

    def _format_phone(self, phone: str) -> str:
        if not phone: return NOT_REGISTERED
        p = re.sub(r'\D', '', phone)
        if len(p) == 11:
            return f"({p[:2]}) {p[2:7]}-{p[7:]}"
        elif len(p) == 10:
            return f"({p[:2]}) {p[2:6]}-{p[6:]}"
        return phone

    def _get_month_name(self, month: int) -> str:
        months = {
            1: "JANEIRO", 2: "FEVEREIRO", 3: "MARÇO", 4: "ABRIL",
            5: "MAIO", 6: "JUNHO", 7: "JULHO", 8: "AGOSTO",
            9: "SETEMBRO", 10: "OUTUBRO", 11: "NOVEMBRO", 12: "DEZEMBRO"
        }
        return months.get(month, "")

    def _time_str_to_fraction(self, time_str: str) -> float:
        if not time_str or ":" not in time_str:
            return 0.0
        parts = time_str.split(":")
        if len(parts) >= 2:
            try:
                hours = int(parts[0])
                minutes = int(parts[1])
                return (hours + (minutes / 60.0)) / 24.0
            except ValueError:
                return 0.0
        return 0.0

    def _validate_employee_report_period(self, current_user: User, month: int, year: int):
        if not current_user:
            return
        from app.domain.models.enums import UserRole
        if current_user.role in [UserRole.MANAGER, UserRole.MAINTAINER]:
            return
            
        from datetime import datetime
        now = datetime.now()
        current_month = now.month
        current_year = now.year
        
        if current_month == 1:
            prev_month = 12
            prev_year = current_year - 1
        else:
            prev_month = current_month - 1
            prev_year = current_year

        if (year == current_year and month == current_month) or \
           (year == prev_year and month == prev_month):
            return
            
        from fastapi import HTTPException
        raise HTTPException(
            status_code=403,
            detail="Funcionários só podem gerar relatório Excel do mês atual ou do mês anterior."
        )

    def generate_excel_report(self, db: Session, month: int, year: int, employee_ids: list[int] | None = None,
                              current_user: User | None = None) -> BytesIO:
        if current_user:
            self._validate_employee_report_period(current_user, month, year)
            
        query = db.query(User).options(joinedload(User.historical_schedules))
        query = report_service._apply_employee_filters(query, employee_ids)
        users = query.all()

        company = company_repository.get_current(db)
        
        logo_path = None
        if company and company.logo_path:
            full_logo_path = os.path.join(settings.UPLOAD_DIR, company.logo_path)
            if os.path.exists(full_logo_path):
                logo_path = full_logo_path

        user_reports = []
        for user in users:
            report = report_service.get_advanced_user_report(db, user.id, month, year, current_user)
            if report and report.summary.total_worked_minutes > 0:
                user_reports.append((user, report))

        wb = Workbook()
        
        self._build_summary_sheet(wb, month, year, user_reports, company, logo_path)

        for user, report in user_reports:
            self._build_employee_sheet(wb, user, report, month, year, company, logo_path)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
        
    def _apply_key_value(self, ws, row, start_col, key_text, key_width, val_text, val_width, borders=False):
        key_end = start_col + key_width - 1
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=key_end)
        c_key = ws.cell(row=row, column=start_col)
        c_key.value = key_text
        c_key.font = self.font_key
        c_key.alignment = self.align_right
        
        val_start = key_end + 1
        val_end = val_start + val_width - 1
        ws.merge_cells(start_row=row, start_column=val_start, end_row=row, end_column=val_end)
        c_val = ws.cell(row=row, column=val_start)
        c_val.value = val_text
        c_val.font = self.font_val
        c_val.alignment = self.align_left

        if borders:
            total_start = start_col
            total_end = val_end
            for c_idx in range(total_start, total_end + 1):
                if c_idx == total_start:
                    ws.cell(row=row, column=c_idx).border = self.border_left_top_bottom
                elif c_idx == total_end:
                    ws.cell(row=row, column=c_idx).border = self.border_right_top_bottom
                else:
                    ws.cell(row=row, column=c_idx).border = self.border_top_bottom

    def _insert_header(self, ws, company, logo_path: str | None):
        company_name = company.name if company else "Empresa Não Cadastrada"
        company_cnpj = self._format_cnpj(company.cnpj if company else "")
        company_phone = self._format_phone(company.phone if company else "")
        company_address = company.address if company else NOT_REGISTERED

        font_large = InlineFont(sz=16, b=True, rFont=FONT_NAME, color="000000")
        font_cnpj_key = InlineFont(sz=11, b=True, rFont=FONT_NAME, color="000000")
        font_cnpj_val = InlineFont(sz=11, b=False, rFont=FONT_NAME, color="000000")
        
        rt = CellRichText(
            TextBlock(font_large, f"{company_name}\n"),
            TextBlock(font_cnpj_key, "CNPJ: "),
            TextBlock(font_cnpj_val, company_cnpj)
        )

        ws.append([""])
        ws.append([""])
        ws.append([""])
        row_start = ws.max_row - 2
        row_mid = row_start + 1
        row_end = ws.max_row
        
        ws.merge_cells(start_row=row_start, start_column=1, end_row=row_end, end_column=self.MAX_COLS)
        c1 = ws.cell(row=row_start, column=1)
        c1.value = rt
        c1.alignment = self.align_center
        
        for r_idx in range(row_start, row_end + 1):
            for c_idx in range(1, self.MAX_COLS + 1):
                ws.cell(row=r_idx, column=c_idx).border = self.border_standard
                ws.cell(row=r_idx, column=c_idx).fill = self.fill_section_title
                
        ws.row_dimensions[row_start].height = 12
        ws.row_dimensions[row_mid].height = 51
        ws.row_dimensions[row_end].height = 12

        ws.append([""])
        row_tel = ws.max_row
        self._apply_key_value(ws, row_tel, start_col=1, key_text="Telefone:", key_width=4, val_text=company_phone, val_width=8, borders=True)
        self._apply_key_value(ws, row_tel, start_col=13, key_text="Endereço:", key_width=3, val_text=company_address, val_width=9, borders=True)
        ws.row_dimensions[row_tel].height = 20

        if logo_path:
            try:
                img = OpenpyxlImage(logo_path)
                img.width = 70
                img.height = 70
                ws.add_image(img, f'A{row_mid}')
            except (OSError, ValueError):
                pass
        ws.append([""])

    def _append_notes(self, ws):
        ws.append([""])
        ws.append(["* Nota: O formato de tempo exibido é HH:MM (Horas:Minutos)."])
        note_row1 = ws.max_row
        ws.merge_cells(start_row=note_row1, start_column=1, end_row=note_row1, end_column=self.MAX_COLS)
        ws.cell(row=note_row1, column=1).font = self.font_italic

        ws.append(["  Exemplo: 10:20 representa exatamente 10 horas e 20 minutos contabilizados."])
        note_row2 = ws.max_row
        ws.merge_cells(start_row=note_row2, start_column=1, end_row=note_row2, end_column=self.MAX_COLS)
        ws.cell(row=note_row2, column=1).font = self.font_italic
        ws.append([""])

    def _merge_for_table(self, ws, row, merges: list[int], texts: list[any], font, alignment, fill=None, borders=True):
        col = 1
        for i, width in enumerate(merges):
            end_col = col + width - 1
            if width > 1:
                ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
            
            for c_idx in range(col, end_col + 1):
                cell = ws.cell(row=row, column=c_idx)
                if borders:
                    cell.border = self.border_standard
                if fill:
                    cell.fill = fill
            
            c = ws.cell(row=row, column=col)
            c.value = texts[i]
            if font:
                c.font = font
            if alignment:
                c.alignment = alignment
            
            col += width
            
    def _build_summary_sheet(self, wb, month, year, user_reports, company, logo_path):
        ws_summary = wb.active
        ws_summary.title = "Resumo"
        self._set_columns_width(ws_summary)
        self._insert_header(ws_summary, company, logo_path)

        subtitle = f"Resumo de Gestão - {self._get_month_name(month).capitalize()} de {year}"
        ws_summary.append([subtitle])
        title_row = ws_summary.max_row
        ws_summary.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=self.MAX_COLS)
        c = ws_summary.cell(row=title_row, column=1)
        c.font = self.font_subtitle
        c.alignment = self.align_center
        for c_idx in range(1, self.MAX_COLS + 1):
            ws_summary.cell(row=title_row, column=c_idx).border = self.border_standard
            ws_summary.cell(row=title_row, column=c_idx).fill = self.fill_section_title
        ws_summary.append([""])

        merges = [14, 5, 5]
        ws_summary.append([""])
        header_row = ws_summary.max_row
        self._merge_for_table(ws_summary, header_row, merges, ["Nome do Colaborador", "Dias Trabalhados", "Horas Trabalhadas"], self.header_font, self.align_center, fill=self.header_fill)

        for user, report in user_reports:
            sum_data = report.summary

            total_real = sum_data.total_worked_minutes / 1440.0

            ws_summary.append([""])
            row = ws_summary.max_row
            texts = [
                sum_data.user_name,
                sum_data.days_worked,
                total_real
            ]
            self._merge_for_table(ws_summary, row, merges, texts, self.font_regular, self.align_center)
            ws_summary.cell(row=row, column=1).alignment = self.align_left
            ws_summary.cell(row=row, column=20).number_format = TIME_FORMAT

        self._append_notes(ws_summary)

    def _build_employee_sheet(self, wb, user, report, month, year, company, logo_path):
        short_name = format_short_name(user.name)
        
        ws_det = wb.create_sheet(title=short_name[:31])
        self._set_columns_width(ws_det)

        self._insert_header(ws_det, company, logo_path)

        title_text = f"Folha de Ponto - {self._get_month_name(month).capitalize()} de {year}"
        ws_det.append([title_text])
        title_row = ws_det.max_row
        ws_det.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=self.MAX_COLS)
        c = ws_det.cell(row=title_row, column=1)
        c.font = self.font_subtitle
        c.alignment = self.align_center
        for c_idx in range(1, self.MAX_COLS + 1):
            ws_det.cell(row=title_row, column=c_idx).border = self.border_standard
            ws_det.cell(row=title_row, column=c_idx).fill = self.fill_section_title
        ws_det.append([""])

        ws_det.append(["Dados do Funcionário"])
        func_title_row = ws_det.max_row
        ws_det.merge_cells(start_row=func_title_row, start_column=1, end_row=func_title_row, end_column=self.MAX_COLS)
        c2 = ws_det.cell(row=func_title_row, column=1)
        c2.font = self.font_bold
        c2.alignment = self.align_left
        for c_idx in range(1, self.MAX_COLS + 1):
            ws_det.cell(row=func_title_row, column=c_idx).border = self.border_standard
            ws_det.cell(row=func_title_row, column=c_idx).fill = self.fill_section_title

        user_cpf = user.cpf or NOT_REGISTERED
        user_pis = user.pis or NOT_REGISTERED
        user_telefone = self._format_phone(user.phone) if hasattr(user, 'phone') and user.phone else NOT_REGISTERED
        user_endereco = user.endereco or NOT_REGISTERED
        
        ws_det.append([""])
        info_row1 = ws_det.max_row
        self._apply_key_value(ws_det, info_row1, start_col=1, key_text="Nome:", key_width=2, val_text=user.name, val_width=8, borders=True)
        self._apply_key_value(ws_det, info_row1, start_col=11, key_text="CPF:", key_width=2, val_text=user_cpf, val_width=5, borders=True)
        self._apply_key_value(ws_det, info_row1, start_col=18, key_text="PIS:", key_width=2, val_text=user_pis, val_width=5, borders=True)
        
        ws_det.append([""])
        info_row2 = ws_det.max_row
        self._apply_key_value(ws_det, info_row2, start_col=1, key_text="Telefone:", key_width=3, val_text=user_telefone, val_width=7, borders=True)
        self._apply_key_value(ws_det, info_row2, start_col=11, key_text="Endereço:", key_width=3, val_text=user_endereco, val_width=11, borders=True)
        ws_det.append([""])

        merges = [3, 3, 3, 8, 2, 2, 3]
        headers_det = ["Data", "Dia Semana", "Status", "Registros", "Trab. Bruto", "Extra Não Aut.", "Trab. Líquido"]
        
        ws_det.append([""])
        header_row = ws_det.max_row
        self._merge_for_table(ws_det, header_row, merges, headers_det, self.header_font, self.align_center, fill=self.header_fill)

        total_trab_bruto = 0.0
        total_extra = 0.0
        total_trab_real = 0.0

        for day in report.daily_details:
            trab_bruto, extra, trab_liquido = self._build_day_row(ws_det, day, merges)
            total_trab_bruto += trab_bruto
            total_extra += extra
            total_trab_real += trab_liquido

        ws_det.append([""])
        last_row = ws_det.max_row
        texts = ["TOTAIS", "", "", "", total_trab_bruto, total_extra, total_trab_real]
        self._merge_for_table(ws_det, last_row, merges, texts, self.font_bold, self.align_center)
        ws_det.cell(row=last_row, column=18).number_format = TIME_FORMAT
        ws_det.cell(row=last_row, column=20).number_format = TIME_FORMAT
        ws_det.cell(row=last_row, column=22).number_format = TIME_FORMAT

        self._append_notes(ws_det)

    def _build_day_row(self, ws_det, day, merges) -> tuple:
        punches_str = " | ".join(day.punches)
        if day.is_holiday:
            holiday_label = day.holiday_name or "Feriado"
            if not punches_str:
                punches_str = holiday_label
            else:
                punches_str = f"{holiday_label} ({punches_str})"

        ws_det.append([""])
        last_row = ws_det.max_row

        trab_liquido = self._time_str_to_fraction(day.worked_time)
        extra_nao_aut = self._time_str_to_fraction(getattr(day, 'unapproved_extra_time', '00:00') or '00:00')
        trab_bruto = trab_liquido + extra_nao_aut

        texts = [
            day.date.strftime("%d/%m/%Y"),
            day.day_name,
            day.status,
            punches_str,
            trab_bruto,
            extra_nao_aut,
            trab_liquido
        ]

        fill_to_apply = None
        font_to_apply = self.font_regular
        if day.is_holiday:
            fill_to_apply = self.fill_holiday
            font_to_apply = self.font_holiday
        elif day.is_weekend:
            fill_to_apply = self.fill_weekend
            font_to_apply = self.font_weekend
        if "Falta" in day.status:
            fill_to_apply = self.fill_absence
            font_to_apply = self.font_absence
        elif "Atestado" in day.status or "Abonado" in day.status:
            fill_to_apply = self.fill_excused
            font_to_apply = self.font_excused

        self._merge_for_table(ws_det, last_row, merges, texts, self.font_regular, self.align_center, fill=fill_to_apply)

        ws_det.cell(row=last_row, column=7).font = font_to_apply

        ws_det.cell(row=last_row, column=18).number_format = TIME_FORMAT
        ws_det.cell(row=last_row, column=20).number_format = TIME_FORMAT
        ws_det.cell(row=last_row, column=22).number_format = TIME_FORMAT

        return trab_bruto, extra_nao_aut, trab_liquido

excel_service = ExcelService()
