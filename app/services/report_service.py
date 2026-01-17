import locale
import pytz
from calendar import monthrange
from datetime import date, timedelta, datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session
from typing import List

from app.core.config import settings
from app.domain.models.enums import RecordType, UserRole, AdjustmentType
from app.domain.models.user import User
from app.repositories.adjustment_repository import adjustment_repository
from app.repositories.holiday_repository import holiday_repository
from app.repositories.time_record_repository import time_record_repository
from app.repositories.user_repository import user_repository
from app.schemas.report import (
    MonthlyReportResponse, UserPayrollSummary, AdvancedUserReportResponse,
    DailyReportItem, DashboardMetricsResponse
)

try:
    locale.setlocale(locale.LC_TIME, 'pt_BR.utf8')
except:
    pass


class ReportService:
    def _get_month_range(self, month: int, year: int):
        start_date = date(year, month, 1)
        _, last_day = monthrange(year, month)
        end_date = date(year, month, last_day)
        return start_date, end_date

    def _get_day_name(self, dt: date) -> str:
        days = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"]
        return days[dt.weekday()]

    def _format_duration(self, total_seconds: float) -> str:
        total_minutes = int(round(total_seconds / 60))
        hours = total_minutes // 60
        minutes = total_minutes % 60
        return f"{hours:02d}:{minutes:02d}"

    def get_dashboard_metrics(self, db: Session) -> DashboardMetricsResponse:
        tz = pytz.timezone(settings.TIMEZONE)
        today = datetime.now(tz).date()

        active_users = db.query(User).filter(
            User.is_active == True,
            User.role == UserRole.EMPLOYEE
        ).count()

        pending = adjustment_repository.count_pending(db)

        today_start = tz.localize(datetime.combine(today, datetime.min.time()))
        today_end = tz.localize(datetime.combine(today, datetime.max.time()))
        present = time_record_repository.count_unique_users_in_range(db, today_start, today_end)

        return DashboardMetricsResponse(
            total_active_employees=active_users,
            pending_adjustments=pending,
            employees_present_today=present,
            date=today
        )

    def get_advanced_user_report(self, db: Session, user_id: int, month: int, year: int) -> AdvancedUserReportResponse:
        start_date, end_date = self._get_month_range(month, year)
        user = user_repository.get(db, user_id)
        if not user:
            return None

        has_schedule = bool(user.schedules)

        tz = pytz.timezone(settings.TIMEZONE)
        today_date = datetime.now(tz).date()

        start_dt = tz.localize(datetime.combine(start_date, datetime.min.time()))
        end_dt = tz.localize(datetime.combine(end_date, datetime.max.time()))

        all_records = time_record_repository.get_by_range(db, user_id, start_dt, end_dt)
        holidays = holiday_repository.get_by_month(db, month, year)
        approved_adjustments = adjustment_repository.get_approved_by_range(db, user_id, start_date, end_date)

        daily_details = []

        total_worked_seconds = 0.0
        total_expected_seconds = 0.0

        total_extra_hours = 0.0
        total_missing_hours = 0.0

        days_worked_count = 0
        absences_count = 0

        current = start_date
        while current <= end_date:
            is_future = current > today_date

            day_records = [r for r in all_records if r.record_datetime.date() == current]
            day_records.sort(key=lambda x: x.record_datetime)

            is_holiday = any(h.date == current for h in holidays)

            adjustment_day = next((adj for adj in approved_adjustments
                                   if adj.target_date == current
                                   and adj.adjustment_type in [AdjustmentType.CERTIFICATE, AdjustmentType.WAIVER]),
                                  None)

            is_certificate = adjustment_day is not None and adjustment_day.adjustment_type == AdjustmentType.CERTIFICATE
            is_waiver = adjustment_day is not None and adjustment_day.adjustment_type == AdjustmentType.WAIVER

            adj_id = adjustment_day.id if adjustment_day else None

            is_excused = is_certificate or is_waiver

            weekday = current.weekday()
            is_weekend = weekday >= 5

            expected_seconds = 0.0
            if has_schedule and not is_holiday and not is_future:
                schedule = next((s for s in user.schedules if s.day_of_week == weekday), None)
                if schedule:
                    expected_seconds = schedule.daily_hours * 3600

            entries = []
            exits = []
            punches = []
            worked_seconds = 0.0
            entry_time = None

            for rec in day_records:
                time_str = rec.record_datetime.strftime("%H:%M")
                suffix = "(E)" if rec.record_type == RecordType.ENTRY else "(S)"
                punches.append(f"{time_str} {suffix}")

                if rec.record_type == RecordType.ENTRY:
                    entries.append(time_str)
                    entry_time = rec.record_datetime
                elif rec.record_type == RecordType.EXIT:
                    exits.append(time_str)
                    if entry_time:
                        delta = rec.record_datetime - entry_time
                        seconds = delta.total_seconds()
                        if seconds <= 86400:
                            worked_seconds += seconds
                        entry_time = None

            waiver_credit = 0.0
            if is_excused:
                if adjustment_day.amount_hours and adjustment_day.amount_hours > 0:
                    waiver_credit = adjustment_day.amount_hours * 3600
                else:
                    if expected_seconds > 0 and worked_seconds < expected_seconds:
                        waiver_credit = expected_seconds - worked_seconds
                worked_seconds += waiver_credit

            if worked_seconds > 0:
                days_worked_count += 1

            if worked_seconds == 0 and expected_seconds > 0 and not is_weekend and not is_holiday and not is_excused and not is_future:
                absences_count += 1

            day_worked_hours = worked_seconds / 3600.0
            day_expected_hours = expected_seconds / 3600.0

            day_balance = 0.0
            if has_schedule:
                day_balance = day_worked_hours - day_expected_hours

            day_extra = day_balance if day_balance > 0 else 0.0
            day_missing = abs(day_balance) if day_balance < 0 else 0.0

            total_worked_seconds += worked_seconds
            total_expected_seconds += expected_seconds

            total_extra_hours += day_extra
            total_missing_hours += day_missing

            status = "Normal"
            if is_future:
                status = ""
            elif is_waiver:
                formatted_waiver = self._format_duration(waiver_credit)
                status = f"Abonado ({formatted_waiver})"
            elif is_certificate:
                formatted_waiver = self._format_duration(waiver_credit)
                status = f"Atestado ({formatted_waiver})"
            elif is_holiday:
                status = "Feriado"
            elif is_weekend:
                status = "Fim de Semana"
            elif worked_seconds == 0 and expected_seconds > 0:
                status = "Falta"
            elif not has_schedule and worked_seconds == 0:
                status = "-"

            worked_minutes_int = int(round(worked_seconds / 60))

            daily_details.append(DailyReportItem(
                date=current,
                day_name=self._get_day_name(current),
                is_holiday=is_holiday,
                is_weekend=is_weekend,
                status=status,
                entries=entries,
                exits=exits,
                punches=punches,

                adjustment_id=adj_id,

                worked_hours=round(day_worked_hours, 2),
                expected_hours=round(day_expected_hours, 2),
                balance_hours=round(day_balance, 2),
                extra_hours=round(day_extra, 2),
                missing_hours=round(day_missing, 2),

                worked_minutes=worked_minutes_int,
                worked_time=self._format_duration(worked_seconds),
                expected_time=self._format_duration(expected_seconds)
            ))

            current += timedelta(days=1)

        summary = UserPayrollSummary(
            user_id=user.id,
            user_name=user.name,
            total_worked_time=self._format_duration(total_worked_seconds),
            total_expected_time=self._format_duration(total_expected_seconds),

            total_worked_minutes=int(round(total_worked_seconds / 60)),
            total_expected_minutes=int(round(total_expected_seconds / 60)),

            days_worked=days_worked_count,
            absences=absences_count,

            total_worked_hours=round(total_worked_seconds / 3600.0, 2),
            total_expected_hours=round(total_expected_seconds / 3600.0, 2),
            total_extra_hours=round(total_extra_hours, 2),
            total_missing_hours=round(total_missing_hours, 2),
            final_balance=round(total_extra_hours - total_missing_hours, 2)
        )

        return AdvancedUserReportResponse(summary=summary, daily_details=daily_details)

    def get_monthly_summary(self, db: Session, month: int, year: int,
                            employee_ids: List[int] = None) -> MonthlyReportResponse:
        query = db.query(User).filter(
            User.is_active == True,
            User.role == UserRole.EMPLOYEE
        )
        if employee_ids:
            query = query.filter(User.id.in_(employee_ids))
        users = query.all()
        payroll_data = []
        for user in users:
            report = self.get_advanced_user_report(db, user.id, month, year)
            if report:
                payroll_data.append(report.summary)
        return MonthlyReportResponse(month=month, year=year, payroll_data=payroll_data)

    def generate_excel_report(self, db: Session, month: int, year: int, employee_ids: List[int] = None) -> BytesIO:
        query = db.query(User).filter(User.is_active == True, User.role == UserRole.EMPLOYEE)
        if employee_ids:
            query = query.filter(User.id.in_(employee_ids))
        users = query.all()

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Resumo Folha"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="003366", end_color="003366",
                                  fill_type="solid")
        border_style = Side(style='thin', color="000000")
        border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

        red_font = Font(color="FF0000", bold=True)
        green_font = Font(color="008000", bold=True)
        blue_font = Font(color="0000FF", bold=True)
        weekend_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        headers_sum = ["Nome do Colaborador", "Dias Trabalhados", "Faltas", "Horas Previstas", "Horas Trabalhadas"]
        ws_summary.append(headers_sum)

        for col_num, header in enumerate(headers_sum, 1):
            cell = ws_summary.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        for user in users:
            report = self.get_advanced_user_report(db, user.id, month, year)
            sum_data = report.summary
            ws_summary.append([
                sum_data.user_name,
                sum_data.days_worked,
                sum_data.absences,
                sum_data.total_expected_time,
                sum_data.total_worked_time
            ])

            last_row = ws_summary.max_row
            for col in range(1, 6):
                ws_summary.cell(row=last_row, column=col).border = border

        for col in ws_summary.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws_summary.column_dimensions[column].width = max_length + 3

        for user in users:
            report = self.get_advanced_user_report(db, user.id, month, year)
            if not report: continue

            sheet_name = f"{user.id}-{user.name.split()[0]}"[:30]
            ws_det = wb.create_sheet(title=sheet_name)

            ws_det.merge_cells('A1:G1')
            title_cell = ws_det['A1']
            title_cell.value = f"Folha de Ponto: {user.name} - {month}/{year}"
            title_cell.font = Font(size=14, bold=True)
            title_cell.alignment = Alignment(horizontal='center')

            headers_det = ["Data", "Dia Semana", "Status", "Registros", "Trabalhado (Min)", "Trabalhado (Tempo)",
                           "Previsto (Tempo)"]
            ws_det.append(headers_det)

            for col_num, header in enumerate(headers_det, 1):
                cell = ws_det.cell(row=2, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')

            for day in report.daily_details:
                punches_str = " | ".join(day.punches)

                row_data = [
                    day.date.strftime("%d/%m/%Y"),
                    day.day_name,
                    day.status,
                    punches_str,
                    day.worked_minutes,
                    day.worked_time,
                    day.expected_time
                ]
                ws_det.append(row_data)
                last_row = ws_det.max_row

                status_cell = ws_det.cell(row=last_row, column=3)

                if "Falta" in day.status:
                    status_cell.font = red_font
                elif "Atestado" in day.status or "Abonado" in day.status:
                    status_cell.font = green_font
                elif "Feriado" in day.status:
                    status_cell.font = blue_font

                if day.is_weekend:
                    for col in range(1, 8):
                        ws_det.cell(row=last_row, column=col).fill = weekend_fill

                for col in range(1, 8):
                    ws_det.cell(row=last_row, column=col).border = border

            ws_det.append([])
            ws_det.append(["TOTAIS", "", "", "",
                           report.summary.total_worked_minutes,
                           report.summary.total_worked_time,
                           report.summary.total_expected_time])

            last_row = ws_det.max_row
            for col in range(1, 8):
                cell = ws_det.cell(row=last_row, column=col)
                cell.font = Font(bold=True)
                cell.border = border

            ws_det.column_dimensions['A'].width = 12
            ws_det.column_dimensions['B'].width = 15
            ws_det.column_dimensions['C'].width = 20
            ws_det.column_dimensions['D'].width = 40
            ws_det.column_dimensions['E'].width = 15
            ws_det.column_dimensions['F'].width = 20
            ws_det.column_dimensions['G'].width = 20

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output


report_service = ReportService()
