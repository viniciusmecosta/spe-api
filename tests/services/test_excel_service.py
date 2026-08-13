import os
import pytest
import sys
from datetime import datetime
from fastapi import HTTPException
from io import BytesIO
from openpyxl import Workbook
from unittest.mock import MagicMock, patch

from app.domain.models.enums import UserRole
from app.domain.models.user import User
from app.services.excel_service import ExcelService


@pytest.fixture
def excel_service():
    return ExcelService()

@pytest.fixture
def mock_user():
    user = MagicMock(spec=User)
    user.id = 1
    user.name = 'Test User'
    user.cpf = '12345678901'
    user.pis = '12345678901'
    user.phone = '11987654321'
    user.endereco = 'Test Address'
    user.role = UserRole.EMPLOYEE
    return user

def test_setup_styles(excel_service):
    assert excel_service.font_regular.name == 'Times New Roman'

def test_set_columns_width(excel_service):
    wb = Workbook()
    ws = wb.active
    excel_service._set_columns_width(ws)
    assert ws.column_dimensions['A'].width == 5

def test_format_cnpj(excel_service):
    assert excel_service._format_cnpj('') == 'Não registrado'
    assert excel_service._format_cnpj('12345678901234') == '12.345.678/9012-34'
    assert excel_service._format_cnpj('123') == '123'

def test_format_phone(excel_service):
    assert excel_service._format_phone('') == 'Não registrado'
    assert excel_service._format_phone('11987654321') == '(11) 98765-4321'
    assert excel_service._format_phone('1187654321') == '(11) 8765-4321'
    assert excel_service._format_phone('118765432') == '118765432'

def test_get_month_name(excel_service):
    assert excel_service._get_month_name(1) == 'JANEIRO'
    assert excel_service._get_month_name(13) == ''

def test_time_str_to_fraction(excel_service):
    assert excel_service._time_str_to_fraction(None) == 0.0
    assert excel_service._time_str_to_fraction('10') == 0.0
    assert excel_service._time_str_to_fraction('invalid:time') == 0.0
    assert excel_service._time_str_to_fraction('aa:bb') == 0.0
    assert excel_service._time_str_to_fraction('10:30') == (10 + 30 / 60.0) / 24.0
    assert excel_service._time_str_to_fraction(':') == 0.0

class MockDatetimeMay:

    class datetime:

        @classmethod
        def now(cls):

            class D:
                month = 5
                year = 2023
            return D()

def test_validate_employee_report_period(excel_service, mock_user, monkeypatch):
    excel_service._validate_employee_report_period(None, 5, 2023)
    manager_user = MagicMock(spec=User)
    manager_user.role = UserRole.MANAGER
    excel_service._validate_employee_report_period(manager_user, 1, 2020)
    monkeypatch.setitem(sys.modules, 'datetime', MockDatetimeMay)
    excel_service._validate_employee_report_period(mock_user, 5, 2023)
    excel_service._validate_employee_report_period(mock_user, 4, 2023)
    with pytest.raises(HTTPException):
        excel_service._validate_employee_report_period(mock_user, 3, 2023)

class MockDatetimeJan:

    class datetime:

        @classmethod
        def now(cls):

            class D:
                month = 1
                year = 2023
            return D()

def test_validate_employee_report_period_january(excel_service, mock_user, monkeypatch):
    monkeypatch.setitem(sys.modules, 'datetime', MockDatetimeJan)
    excel_service._validate_employee_report_period(mock_user, 1, 2023)
    excel_service._validate_employee_report_period(mock_user, 12, 2022)
    with pytest.raises(HTTPException):
        excel_service._validate_employee_report_period(mock_user, 11, 2022)

@patch.object(ExcelService, '_validate_employee_report_period')
@patch('app.services.excel_service.report_service')
@patch('app.services.excel_service.company_repository')
@patch('os.path.exists')
def test_generate_excel_report(mock_exists, mock_company_repo, mock_report_service, mock_validate, excel_service, db_session_mock, mock_user):
    mock_exists.return_value = True
    mock_company = MagicMock()
    mock_company.logo_path = 'logo.png'
    mock_company.cnpj = '12345678901234'
    mock_company.phone = '11987654321'
    mock_company.address = 'Address'
    mock_company.name = 'Company'
    mock_company_repo.get_current.return_value = mock_company
    query_mock = MagicMock()
    query_mock.options.return_value = query_mock
    query_mock.filter.return_value = query_mock
    def mock_query_side_effect(model):
        if getattr(model, '__name__', '') == 'User':
            query_mock.all.return_value = [mock_user]
            return query_mock
        else:
            other_mock = MagicMock()
            other_mock.filter.return_value = other_mock
            other_mock.all.return_value = []
            return other_mock
    db_session_mock.query.side_effect = mock_query_side_effect
    mock_report_service._apply_employee_filters.return_value = query_mock
    mock_report_service._get_month_range.return_value = (datetime(2023, 5, 1).date(), datetime(2023, 5, 31).date())
    mock_report_service._get_month_range.return_value = (datetime(2023, 5, 1).date(), datetime(2023, 5, 31).date())
    mock_report = MagicMock()
    mock_report.summary.total_worked_minutes = 100
    mock_report.summary.user_name = 'Test User'
    mock_report.summary.days_worked = 1
    mock_day = MagicMock()
    mock_day.worked_time = '08:00'
    mock_day.unapproved_extra_time = '01:00'
    mock_day.is_holiday = False
    mock_day.is_weekend = False
    mock_day.status = 'Normal'
    mock_day.punches = ['08:00', '12:00']
    mock_day.date = datetime(2023, 5, 1)
    mock_day.day_name = 'Segunda'
    mock_report.daily_details = [mock_day]
    mock_report_service.get_advanced_user_report.return_value = mock_report
    output = excel_service.generate_excel_report(db_session_mock, 5, 2023, [1], mock_user)
    assert isinstance(output, BytesIO)
    mock_validate.assert_called_once_with(mock_user, 5, 2023)

@patch('app.services.excel_service.report_service')
@patch('app.services.excel_service.company_repository')
def test_generate_excel_report_no_logo(mock_company_repo, mock_report_service, excel_service, db_session_mock, mock_user):
    mock_company = MagicMock()
    mock_company.logo_path = None
    mock_company.cnpj = '12345678901234'
    mock_company.phone = '11987654321'
    mock_company.address = 'Address'
    mock_company.name = 'Company'
    mock_company_repo.get_current.return_value = mock_company
    query_mock = MagicMock()
    query_mock.options.return_value = query_mock
    query_mock.filter.return_value = query_mock
    def mock_query_side_effect(model):
        if getattr(model, '__name__', '') == 'User':
            query_mock.all.return_value = [mock_user]
            return query_mock
        else:
            other_mock = MagicMock()
            other_mock.filter.return_value = other_mock
            other_mock.all.return_value = []
            return other_mock
    db_session_mock.query.side_effect = mock_query_side_effect
    mock_report_service._apply_employee_filters.return_value = query_mock
    mock_report_service._get_month_range.return_value = (datetime(2023, 5, 1).date(), datetime(2023, 5, 31).date())
    mock_report = MagicMock()
    mock_report.summary.total_worked_minutes = 100
    mock_report.summary.user_name = 'Test User'
    mock_report.summary.days_worked = 1
    mock_day = MagicMock()
    mock_day.worked_time = '08:00'
    mock_day.unapproved_extra_time = None
    mock_day.is_holiday = True
    mock_day.holiday_name = 'Dia do Trabalho'
    mock_day.is_weekend = False
    mock_day.status = 'Feriado'
    mock_day.punches = []
    mock_day.date = datetime(2023, 5, 1)
    mock_day.day_name = 'Segunda'
    mock_report.daily_details = [mock_day]
    mock_report_service.get_advanced_user_report.return_value = mock_report
    output = excel_service.generate_excel_report(db_session_mock, 5, 2023, [1])
    assert isinstance(output, BytesIO)

def test_apply_key_value(excel_service):
    wb = Workbook()
    ws = wb.active
    excel_service._apply_key_value(ws, 1, 1, 'Key', 2, 'Value', 2, borders=True)
    assert ws.cell(row=1, column=1).value == 'Key'
    assert ws.cell(row=1, column=3).value == 'Value'
    excel_service._apply_key_value(ws, 2, 1, 'Key', 2, 'Value', 2, borders=False)

@patch('app.services.excel_service.OpenpyxlImage')
def test_insert_header(mock_image, excel_service):
    wb = Workbook()
    ws = wb.active
    mock_company = MagicMock()
    mock_company.cnpj = '12345678901234'
    mock_company.phone = '11987654321'
    mock_company.address = 'Address'
    excel_service._insert_header(ws, mock_company, 'logo.png')
    assert ws.max_row > 1

@patch('app.services.excel_service.OpenpyxlImage')
def test_insert_header_no_company(mock_image, excel_service):
    wb = Workbook()
    ws = wb.active
    excel_service._insert_header(ws, None, None)
    assert ws.max_row > 1

@patch('app.services.excel_service.OpenpyxlImage')
def test_insert_header_image_error(mock_image, excel_service):
    mock_image.side_effect = ValueError('Invalid image')
    wb = Workbook()
    ws = wb.active
    mock_company = MagicMock()
    mock_company.cnpj = '12345678901234'
    mock_company.phone = '11987654321'
    mock_company.address = 'Address'
    excel_service._insert_header(ws, mock_company, 'logo.png')
    mock_image.side_effect = OSError('No file')
    excel_service._insert_header(ws, mock_company, 'logo.png')

def test_append_notes(excel_service):
    wb = Workbook()
    ws = wb.active
    excel_service._append_notes(ws)
    assert ws.max_row >= 4

def test_merge_for_table(excel_service):
    wb = Workbook()
    ws = wb.active
    excel_service._merge_for_table(ws, 1, [2, 2], ['Text1', 'Text2'], excel_service.font_regular, excel_service.align_center, fill=excel_service.fill_holiday, borders=True)
    excel_service._merge_for_table(ws, 2, [1], ['Text'], None, None, fill=None, borders=False)

def test_build_day_row(excel_service):
    wb = Workbook()
    ws = wb.active
    mock_day = MagicMock()
    mock_day.worked_time = '08:00'
    mock_day.unapproved_extra_time = '01:00'
    mock_day.is_holiday = False
    mock_day.is_weekend = True
    mock_day.status = 'Falta'
    mock_day.punches = ['08:00', '12:00']
    mock_day.date = datetime(2023, 5, 1)
    mock_day.day_name = 'Sábado'
    excel_service._build_day_row(ws, mock_day, [1, 1, 1, 1, 1, 1])
    mock_day.is_holiday = True
    mock_day.is_weekend = False
    mock_day.status = 'Atestado'
    mock_day.punches = ['08:00']
    mock_day.holiday_name = None
    excel_service._build_day_row(ws, mock_day, [1, 1, 1, 1, 1, 1])
    mock_day.status = 'Abonado'
    excel_service._build_day_row(ws, mock_day, [1, 1, 1, 1, 1, 1])
    mock_day.is_holiday = True
    mock_day.punches = []
    mock_day.status = 'Feriado'
    excel_service._build_day_row(ws, mock_day, [1, 1, 1, 1, 1, 1])

def test_build_employee_sheet_no_phone_endereco(excel_service):
    wb = Workbook()
    mock_user = MagicMock(spec=User)
    mock_user.name = 'Test User'
    mock_user.cpf = None
    mock_user.pis = None
    mock_user.phone = None
    mock_user.endereco = None
    mock_report = MagicMock()
    mock_report.daily_details = []
    from datetime import date
    excel_service._build_employee_sheet(wb, mock_user, mock_report, 5, 2023, None, None, date(2023, 5, 1),
                                        date(2023, 5, 31))
    assert 'Tes' in wb.sheetnames[-1]

def test_build_summary_sheet(excel_service):
    wb = Workbook()
    mock_user = MagicMock(spec=User)
    mock_user.name = 'Test User'
    mock_report = MagicMock()
    mock_report.summary.user_name = 'Test User'
    mock_report.summary.days_worked = 1
    mock_report.summary.total_worked_minutes = 480
    mock_day = MagicMock()
    mock_day.worked_time = '08:00'
    mock_day.unapproved_extra_time = '01:00'
    mock_report.daily_details = [mock_day]
    excel_service._build_summary_sheet(wb, 5, 2023, [(mock_user, mock_report)], None, None)
    assert 'Resumo' in wb.sheetnames

def test_build_summary_sheet_bruto_less_extra(excel_service):
    wb = Workbook()
    mock_user = MagicMock(spec=User)
    mock_user.name = 'Test User'
    mock_report = MagicMock()
    mock_report.summary.user_name = 'Test User'
    mock_report.summary.days_worked = 1
    mock_report.summary.total_worked_minutes = 480
    mock_day = MagicMock()
    mock_day.worked_time = '01:00'
    mock_day.unapproved_extra_time = '02:00'
    mock_report.daily_details = [mock_day]
    excel_service._build_summary_sheet(wb, 5, 2023, [(mock_user, mock_report)], None, None)
    assert 'Resumo' in wb.sheetnames
import openpyxl
from app.schemas.report import UserPayrollSummary, DailyReportItem

@patch.object(ExcelService, '_validate_employee_report_period')
@patch('app.services.excel_service.report_service')
@patch('app.services.excel_service.company_repository')
@patch('os.path.exists')
def test_exhaustive_excel_structural_generation(mock_exists, mock_company_repo, mock_report_service, mock_validate, excel_service, db_session_mock, mock_user):
    mock_exists.return_value = False
    mock_company = MagicMock()
    mock_company.logo_path = None
    mock_company.cnpj = '12345678901234'
    mock_company.phone = '11987654321'
    mock_company.address = 'Address'
    mock_company.name = 'Company'
    mock_company_repo.get_current.return_value = mock_company
    query_mock = MagicMock()
    query_mock.options.return_value = query_mock
    query_mock.filter.return_value = query_mock
    def mock_query_side_effect(model):
        if getattr(model, '__name__', '') == 'User':
            query_mock.all.return_value = [mock_user]
            return query_mock
        else:
            other_mock = MagicMock()
            other_mock.filter.return_value = other_mock
            other_mock.all.return_value = []
            return other_mock
    db_session_mock.query.side_effect = mock_query_side_effect
    mock_report_service._apply_employee_filters.return_value = query_mock
    mock_report_service._get_month_range.return_value = (datetime(2023, 10, 1).date(), datetime(2023, 10, 31).date())
    mock_report = MagicMock()
    mock_report.summary = UserPayrollSummary(user_id=1, user_name='Teste Silva', total_worked_time='10:00', total_expected_time='08:00', total_worked_minutes=600, total_expected_minutes=480, days_worked=2, absences=1, total_worked_hours=10.0, total_expected_hours=8.0, total_extra_hours=2.0, total_missing_hours=0.0, final_balance=2.0)
    day1 = DailyReportItem(date=datetime(2023, 10, 10), day_name='Terça', is_holiday=False, is_weekend=False, status='Normal', worked_hours=10.0, expected_hours=8.0, balance_hours=2.0, extra_hours=2.0, missing_hours=0.0, worked_minutes=600, worked_time='10:00', expected_time='08:00', unapproved_extra_time='00:00', punches=['08:00', '18:00'], holiday_name=None, entries=[], exits=[], detailed_punches=None, adjustment_id=None)
    day2 = DailyReportItem(date=datetime(2023, 10, 12), day_name='Quinta', is_holiday=True, is_weekend=False, status='Feriado', worked_hours=0.0, expected_hours=8.0, balance_hours=0.0, extra_hours=0.0, missing_hours=0.0, worked_minutes=0, worked_time='00:00', expected_time='08:00', unapproved_extra_time='00:00', punches=[], holiday_name='Nossa Sra', entries=[], exits=[], detailed_punches=None, adjustment_id=None)
    day3 = DailyReportItem(date=datetime(2023, 10, 13), day_name='Sexta', is_holiday=False, is_weekend=False, status='Falta', worked_hours=0.0, expected_hours=8.0, balance_hours=-8.0, extra_hours=0.0, missing_hours=8.0, worked_minutes=0, worked_time='00:00', expected_time='08:00', unapproved_extra_time='00:00', punches=[], holiday_name=None, entries=[], exits=[], detailed_punches=None, adjustment_id=None)
    mock_report.daily_details = [day1, day2, day3]
    mock_report_service.get_advanced_user_report.return_value = mock_report
    output = excel_service.generate_excel_report(db_session_mock, 10, 2023, [1])
    wb = openpyxl.load_workbook(output)
    assert 'Resumo' in wb.sheetnames, 'Aba Resumo deve existir'
    assert 'Test User' in wb.sheetnames, 'Aba com nome do funcionário deve existir'
    ws_resumo = wb['Resumo']
    company_title = ws_resumo['A1'].value
    assert 'Company' in str(company_title)
    ws_func = wb['Test User']
    data_header_row = None
    for row in ws_func.iter_rows(min_row=1, max_row=20):
        if row[0].value == 'Data':
            data_header_row = row[0].row
            break
    assert data_header_row is not None
    row_day1 = data_header_row + 1
    row_day2 = data_header_row + 2
    row_day3 = data_header_row + 3
    cell_day2 = ws_func.cell(row=row_day2, column=6)
    assert cell_day2.fill.start_color.rgb in ['00FEF3C7', 'FEF3C7', 'FFFEF3C7']
    cell_day3 = ws_func.cell(row=row_day3, column=6)
    assert cell_day3.fill.start_color.rgb in ['00FEE2E2', 'FEE2E2', 'FFFEE2E2']
    assert ws_func.cell(row=data_header_row, column=1).font.bold == True
    assert ws_func.cell(row=row_day1, column=1).border.left.style is not None
    assert ws_func.cell(row=row_day1, column=1).alignment.horizontal is not None

def test_format_day_groups(excel_service):
    assert excel_service._format_day_groups([0, 1, 2, 3, 4]) == "Segunda a Sexta"
    assert excel_service._format_day_groups([0, 2, 4]) == "Segunda, Quarta e Sexta"
    assert excel_service._format_day_groups([0, 1, 2, 4, 5, 6]) == "Segunda a Quarta e Sexta a Domingo"
    assert excel_service._format_day_groups([5]) == "Sábado"
    assert excel_service._format_day_groups([]) == ""

def test_build_work_schedules_section(excel_service):
    from app.domain.models.user import UserWorkScheduleConfig
    from datetime import date, time
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    
    mock_user = MagicMock(spec=User)
    sch1 = UserWorkScheduleConfig(
        day_of_week=0,
        valid_from=date(2023, 5, 1),
        valid_until=date(2023, 5, 14),
        entry_1=time(8, 0), exit_1=time(12, 0), entry_2=time(13, 0), exit_2=time(17, 0)
    )
    sch2 = UserWorkScheduleConfig(
        day_of_week=1,
        valid_from=date(2023, 5, 15),
        valid_until=None,
        entry_1=time(9, 0), exit_1=time(13, 0), entry_2=time(14, 0), exit_2=time(18, 0)
    )
    mock_user.historical_schedules = [sch1, sch2]
    
    excel_service._build_work_schedules_section(ws, mock_user, date(2023, 5, 1), date(2023, 5, 31))
    
    assert ws.max_row > 1
    found_title = False
    found_first_period = False
    found_second_period = False
    for row in ws.iter_rows(values_only=True):
        if row[0] == "Expediente Cadastrado":
            found_title = True
        if row[0] == "Período: 01/05/2023 a 14/05/2023":
            found_first_period = True
        if row[0] == "Período: 15/05/2023 a 31/05/2023":
            found_second_period = True
            
    assert found_title
    assert found_first_period
    assert found_second_period

def test_format_day_groups_two_days(excel_service):
    assert excel_service._format_day_groups([0, 1]) == "Segunda e Terça"

def test_build_work_schedules_section_edge_cases(excel_service):
    from datetime import date, time
    from openpyxl import Workbook
    from app.domain.models.user import UserWorkScheduleConfig
    
    wb = Workbook()
    ws = wb.active
    
    mock_user_no_sched = MagicMock(spec=User)
    mock_user_no_sched.historical_schedules = []
    excel_service._build_work_schedules_section(ws, mock_user_no_sched, date(2023, 5, 1), date(2023, 5, 31))
    
    mock_user_empty_periods = MagicMock(spec=User)
    mock_user_empty_periods.historical_schedules = [
        UserWorkScheduleConfig(day_of_week=0, valid_from=date(2023, 5, 1), valid_until=date(2023, 5, 1))
    ]
    with patch.object(excel_service, "_group_schedules_by_period", return_value=[]):
        excel_service._build_work_schedules_section(ws, mock_user_empty_periods, date(2023, 5, 1), date(2023, 5, 31))

def test_group_schedules_by_period_start_greater_than_end(excel_service):
    from datetime import date
    user = MagicMock()
    user.historical_schedules = []
    transitions = [date(2023, 5, 10), date(2023, 5, 10)]
    periods = excel_service._group_schedules_by_period(user, transitions)
    assert periods == []

def test_write_period_schedules_empty_entries_and_no_grouped(excel_service):
    from openpyxl import Workbook
    from app.domain.models.user import UserWorkScheduleConfig
    from datetime import date
    
    wb = Workbook()
    ws = wb.active
    
    sch_empty = UserWorkScheduleConfig(
        day_of_week=0,
        valid_from=date(2023, 5, 1),
        valid_until=date(2023, 5, 31),
        entry_1=None, exit_1=None, entry_2=None, exit_2=None
    )
    excel_service._write_period_schedules(ws, date(2023, 5, 1), date(2023, 5, 31), [sch_empty], is_single_period=False)
    
    found_no_sched = False
    for row in ws.iter_rows(values_only=True):
        if "Sem expediente cadastrado" in str(row):
            found_no_sched = True
    assert found_no_sched

def test_time_str_to_fraction_fallback(excel_service):
    class CustomStr(str):
        def __contains__(self, item):
            return True
        def split(self, sep=None, maxsplit=-1):
            return ["10"]
    assert excel_service._time_str_to_fraction(CustomStr("10:00")) == 0.0

def test_build_day_row_abono_status(excel_service):
    from openpyxl import Workbook
    from datetime import datetime
    from app.schemas.report import DailyReportItem
    
    wb = Workbook()
    ws = wb.active
    
    day_abono = DailyReportItem(
        date=datetime(2023, 10, 10),
        day_name="Terça",
        is_holiday=False,
        is_weekend=False,
        status="Abono Parcial",
        worked_hours=4.0,
        expected_hours=8.0,
        balance_hours=-4.0,
        extra_hours=0.0,
        missing_hours=4.0,
        worked_minutes=240,
        worked_time="04:00",
        expected_time="08:00",
        unapproved_extra_time="00:00",
        punches=["08:00", "12:00"],
        holiday_name=None,
        entries=[],
        exits=[],
        detailed_punches=None,
        adjustment_id=None
    )
    
    merges = [2, 3, 13, 2, 2, 2]
    excel_service._build_day_row(ws, day_abono, merges)
    last_row = ws.max_row
    assert ws.cell(row=last_row, column=6).fill == excel_service.fill_excused

@patch("app.services.excel_service.company_repository")
@patch("app.services.excel_service.report_service")
def test_generate_excel_report_with_records_and_adjustments(mock_report_service, mock_comp_repo, excel_service, db_session_mock):
    from datetime import datetime, date
    from app.domain.models.time_record import TimeRecord
    from app.domain.models.adjustment import AdjustmentRequest
    from app.schemas.report import AdvancedUserReportResponse, UserPayrollSummary
    
    user = MagicMock(spec=User)
    user.id = 1
    user.name = "Test User Batch"
    user.role = UserRole.EMPLOYEE
    user.historical_schedules = []
    
    query_mock = MagicMock()
    query_mock.options.return_value = query_mock
    query_mock.all.return_value = [user]
    
    rec = MagicMock(spec=TimeRecord)
    rec.user_id = 1
    rec.record_datetime = datetime(2023, 10, 10, 8, 0)
    
    adj = MagicMock(spec=AdjustmentRequest)
    adj.user_id = 1
    adj.target_date = date(2023, 10, 10)
    
    def side_query_filter(model):
        m = MagicMock()
        if model == TimeRecord:
            m.filter.return_value.all.return_value = [rec]
        elif model == AdjustmentRequest:
            m.filter.return_value.all.return_value = [adj]
        else:
            m.options.return_value.all.return_value = [user]
        return m
        
    db_session_mock.query.side_effect = side_query_filter
    mock_comp_repo.get_current.return_value = None
    mock_report_service._get_month_range.return_value = (date(2023, 10, 1), date(2023, 10, 31))
    mock_report_service._apply_employee_filters.side_effect = lambda q, e: q
    
    mock_rep = MagicMock(spec=AdvancedUserReportResponse)
    mock_rep.summary = UserPayrollSummary(
        user_id=1, user_name="Test User Batch", total_worked_time="00:00", total_expected_time="00:00",
        total_worked_minutes=0, total_expected_minutes=0, days_worked=0, absences=0,
        total_worked_hours=0.0, total_expected_hours=0.0, total_extra_hours=0.0, total_missing_hours=0.0, final_balance=0.0
    )
    mock_rep.daily_details = []
    mock_report_service.get_advanced_user_report.return_value = mock_rep
    
    res = excel_service.generate_excel_report(db_session_mock, 10, 2023, [1])
    assert res is not None
def test_generate_excel_report_filters_ignored_records(excel_service, db_session_mock):
    from app.domain.models.user import User
    from unittest.mock import patch, MagicMock
    
    user = MagicMock(spec=User)
    user.id = 1
    user.name = 'Test'
    user.historical_schedules = []
    
    mock_query_tr = MagicMock()
    mock_query_tr.options.return_value = mock_query_tr
    mock_query_tr.filter.return_value = mock_query_tr
    mock_query_tr.all.return_value = []
    
    def mock_query_side_effect(model):
        mock_q = MagicMock()
        mock_q.options.return_value = mock_q
        if getattr(model, '__name__', '') == 'User':
            mock_q.filter.return_value = mock_q
            mock_q.all.return_value = [user]
            return mock_q
        elif getattr(model, '__name__', '') == 'TimeRecord':
            return mock_query_tr
        else:
            mock_q.filter.return_value = mock_q
            mock_q.all.return_value = []
            return mock_q
            
    db_session_mock.query.side_effect = mock_query_side_effect
    
    with patch('app.services.excel_service.company_repository.get_current', return_value=None), \
         patch('app.services.report_service.report_service.get_advanced_user_report', return_value=None):
        excel_service.generate_excel_report(db_session_mock, 1, 2024, [1])
        
    filter_args = mock_query_tr.filter.call_args[0]
    is_ignored_filtered = False
    for arg in filter_args:
        if 'is_ignored' in str(arg) and 'false' in str(arg).lower():
            is_ignored_filtered = True
            break
            
    assert is_ignored_filtered, "O filtro TimeRecord.is_ignored == False deve ser aplicado no ExcelService!"
